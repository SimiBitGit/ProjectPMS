"""
import_etf_universe.py — Import-Script für das Sub-Industry ETF Universe

Liest die Datei 'sub_industry_etf_universe.xlsx' und importiert alle ETFs
in die Datenbank via TickerService (inkl. automatischer GICS-Denormalisierung).

Verwendung:
    python -m src.scripts.import_etf_universe
    python -m src.scripts.import_etf_universe --dry-run
    python -m src.scripts.import_etf_universe --file pfad/zur/datei.xlsx
    python -m src.scripts.import_etf_universe --update   # überschreibt bestehende Ticker

Voraussetzungen:
    1. Datenbank initialisiert: python -m src.database.init_db
    2. GICS-Daten geladen (geschieht automatisch beim ersten Run)
    3. openpyxl installiert: pip install openpyxl
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from typing import Optional

import openpyxl

from src.models.base import get_session
from src.models.metadata import AssetType, GicsLevel
from src.database.gics_repository import GicsRepository
from src.services.ticker_service import (
    TickerService,
    TickerCreateDTO,
    TickerUpdateDTO,
    TickerAlreadyExistsError,
    InvalidGicsCodeError,
    InvalidTickerDataError,
)
from src.utils.logger import get_logger

logger = get_logger(__name__)

# Pfad zur Excel-Datei relativ zum Projekt-Root
DEFAULT_XLSX_PATH = Path("data/imports/sub_industry_etf_universe.xlsx")

# Spalten-Mapping (0-basiert, entspricht den Spalten im Sheet)
COL_SECTOR_NAME  = 0
COL_SUB_IND_NAME = 1
COL_GICS_CODE    = 2
COL_SYMBOL       = 3
COL_ETF_NAME     = 4
COL_EXCHANGE     = 5
COL_COUNTRY      = 6
COL_PROVIDER     = 7
COL_TER          = 8
COL_AUM          = 9
COL_ISIN         = 10
COL_NOTE         = 11

# Mapping Länder → Währung
COUNTRY_CURRENCY = {
    "US": "USD",
    "DE": "EUR",
    "CH": "CHF",
    "UK": "GBP",
    "CA": "CAD",
    "AU": "AUD",
    "JP": "JPY",
    "FR": "EUR",
}

# Mapping Börse → Domizil
EXCHANGE_DOMICILE = {
    "NYSE": "US", "NASDAQ": "US",
    "XETRA": "DE", "SIX": "CH",
    "LSE": "UK", "TSX": "CA", "ASX": "AU",
    "TSE": "JP", "TYO": "JP",
}


@dataclass
class ImportResult:
    created:  int = 0
    updated:  int = 0
    skipped:  int = 0
    errors:   int = 0
    messages: list[str] = field(default_factory=list)

    def summary(self) -> str:
        return (
            f"\n{'='*60}\n"
            f"  Import abgeschlossen\n"
            f"  Erstellt:   {self.created}\n"
            f"  Aktualisiert: {self.updated}\n"
            f"  Übersprungen: {self.skipped}\n"
            f"  Fehler:     {self.errors}\n"
            f"{'='*60}"
        )


def _parse_row(row: tuple) -> Optional[dict]:
    """
    Liest eine Excel-Zeile und gibt ein dict mit den Rohdaten zurück.
    Gibt None zurück wenn die Zeile leer oder eine Sektor-Headerzeile ist.
    """
    values = [cell.value for cell in row]

    # Überspringe leere Zeilen und Sektor-Headerzeilen (Symbol-Spalte leer)
    symbol = values[COL_SYMBOL]
    if not symbol or str(symbol).strip() == "":
        return None

    symbol = str(symbol).strip().upper()

    # Sektor-Trennzeilen haben kein Symbol
    if str(values[COL_SECTOR_NAME] or "").startswith(" ") and not symbol:
        return None

    return {
        "symbol":       symbol,
        "etf_name":     str(values[COL_ETF_NAME] or "").strip(),
        "gics_code":    str(values[COL_GICS_CODE] or "").strip(),
        "sub_ind_name": str(values[COL_SUB_IND_NAME] or "").strip(),
        "exchange":     str(values[COL_EXCHANGE] or "").strip().upper(),
        "country":      str(values[COL_COUNTRY] or "").strip().upper(),
        "provider":     str(values[COL_PROVIDER] or "").strip(),
        "ter":          values[COL_TER],
        "aum":          values[COL_AUM],
        "isin":         str(values[COL_ISIN] or "").strip() or None,
        "note":         str(values[COL_NOTE] or "").strip() or None,
    }


def _build_dto(row_data: dict, gics_repo: GicsRepository) -> TickerCreateDTO:
    """
    Baut ein TickerCreateDTO aus den geparsten Zeilendaten.

    GICS-Logik:
    - Wenn gics_code 8-stellig ist → sub_industry_code setzen (Service denormalisiert auto.)
    - Wenn 6-stellig → industry_code
    - Wenn 4-stellig → industry_group_code
    - Wenn 2-stellig → sector_code
    """
    gics_code = row_data["gics_code"]
    exchange  = row_data["exchange"]
    country   = row_data["country"]

    # TER: Im Sheet als Dezimalzahl (z.B. 0.0035 = 0.35%) oder Prozentwert (0.35)
    ter = None
    if row_data["ter"] is not None:
        ter_val = float(row_data["ter"])
        # Normalisieren: Werte > 0.1 sind bereits in Prozent (z.B. 0.35 → 0.35%)
        # Werte <= 0.1 sind als Dezimal (z.B. 0.0035 → 0.35%)
        if ter_val > 0.1:
            ter = Decimal(str(round(ter_val, 4)))
        else:
            ter = Decimal(str(round(ter_val * 100, 4)))

    aum = None
    if row_data["aum"] is not None:
        aum = Decimal(str(round(float(row_data["aum"]) * 1_000_000_000, 2)))

    # GICS-Felder aufteilen
    kwargs = {}
    if len(gics_code) == 8:
        kwargs["gics_sub_industry_code"] = gics_code
    elif len(gics_code) == 6:
        kwargs["gics_industry_code"]      = gics_code
        kwargs["gics_level"]              = GicsLevel.INDUSTRY
    elif len(gics_code) == 4:
        kwargs["gics_industry_group_code"] = gics_code
        kwargs["gics_level"]               = GicsLevel.INDUSTRY_GROUP
    elif len(gics_code) == 2:
        kwargs["gics_sector_code"] = gics_code
        kwargs["gics_level"]       = GicsLevel.SECTOR

    currency = COUNTRY_CURRENCY.get(country, "USD")
    domicile = EXCHANGE_DOMICILE.get(exchange, country)

    return TickerCreateDTO(
        symbol=row_data["symbol"],
        asset_type=AssetType.ETF,
        name=row_data["etf_name"],
        exchange=exchange,
        currency=currency,
        isin=row_data["isin"],
        etf_provider=row_data["provider"] or None,
        ter_percent=ter,
        aum_usd=aum,
        domicile=domicile,
        **kwargs,
    )


def run_import(
    xlsx_path: Path,
    dry_run: bool = False,
    update_existing: bool = False,
    sheet_name: str = "Sub-Industry ETFs",
) -> ImportResult:
    """
    Hauptfunktion: Liest die Excel-Datei und importiert alle ETFs.

    Args:
        xlsx_path:        Pfad zur Excel-Datei
        dry_run:          Wenn True, nur validieren ohne DB-Schreibzugriff
        update_existing:  Wenn True, bestehende Ticker aktualisieren
        sheet_name:       Name des Sheets in der Excel-Datei
    """
    result = ImportResult()

    if not xlsx_path.exists():
        msg = f"Excel-Datei nicht gefunden: {xlsx_path}"
        logger.error(msg)
        result.messages.append(f"FEHLER: {msg}")
        result.errors += 1
        return result

    logger.info(f"Öffne Excel-Datei: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if sheet_name not in wb.sheetnames:
        # Fallback: erstes Sheet
        ws = wb.active
        logger.warning(f"Sheet '{sheet_name}' nicht gefunden, nutze: {ws.title}")
    else:
        ws = wb[sheet_name]

    session = get_session()
    service = TickerService(session)
    gics_repo = GicsRepository(session)

    # GICS-Seed-Daten sicherstellen
    if not dry_run:
        seeded = gics_repo.seed_gics_data()
        if seeded > 0:
            logger.info(f"GICS-Referenzdaten geladen: {seeded} neue Einträge")

    # Header-Zeile überspringen
    rows = list(ws.iter_rows(min_row=2))
    logger.info(f"Verarbeite {len(rows)} Zeilen...")

    for row in rows:
        row_data = _parse_row(row)

        # Überspringe Leer- und Headerzeilen
        if row_data is None:
            continue

        symbol = row_data["symbol"]

        try:
            dto = _build_dto(row_data, gics_repo)
        except Exception as e:
            msg = f"[{symbol}] DTO-Fehler: {e}"
            logger.warning(msg)
            result.messages.append(f"FEHLER: {msg}")
            result.errors += 1
            continue

        if dry_run:
            logger.info(f"[DRY-RUN] {symbol} — {row_data['etf_name'][:50]} | GICS: {row_data['gics_code']}")
            result.created += 1
            continue

        try:
            service.create_ticker(dto)
            logger.info(f"[ERSTELLT] {symbol} — {row_data['etf_name'][:50]}")
            result.created += 1

        except TickerAlreadyExistsError:
            if update_existing:
                # Update: nur veränderliche Felder
                update_dto = TickerUpdateDTO(
                    name=dto.name,
                    etf_provider=dto.etf_provider,
                    ter_percent=dto.ter_percent,
                    aum_usd=dto.aum_usd,
                    isin=dto.isin,
                    gics_sub_industry_code=dto.gics_sub_industry_code,
                    gics_sector_code=dto.gics_sector_code,
                    gics_industry_group_code=dto.gics_industry_group_code,
                    gics_industry_code=dto.gics_industry_code,
                    gics_level=dto.gics_level,
                )
                try:
                    service.update_ticker(symbol, update_dto)
                    logger.info(f"[AKTUALISIERT] {symbol}")
                    result.updated += 1
                except Exception as e:
                    msg = f"[{symbol}] Update-Fehler: {e}"
                    logger.error(msg)
                    result.messages.append(f"FEHLER: {msg}")
                    result.errors += 1
            else:
                logger.debug(f"[ÜBERSPRUNGEN] {symbol} existiert bereits")
                result.skipped += 1

        except InvalidGicsCodeError as e:
            session.rollback()
            msg = f"[{symbol}] Ungültiger GICS-Code: {e}"
            logger.error(msg)
            result.messages.append(f"FEHLER: {msg}")
            result.errors += 1

        except InvalidTickerDataError as e:
            session.rollback()
            msg = f"[{symbol}] Ungültige Daten: {e}"
            logger.error(msg)
            result.messages.append(f"FEHLER: {msg}")
            result.errors += 1

        except Exception as e:
            session.rollback()
            msg = f"[{symbol}] Unerwarteter Fehler: {e}"
            logger.error(msg)
            result.messages.append(f"FEHLER: {msg}")
            result.errors += 1

    if not dry_run:
        session.close()

    return result


def main():
    parser = argparse.ArgumentParser(
        description="Importiert Sub-Industry ETFs aus Excel in die Portfolio-Datenbank"
    )
    parser.add_argument(
        "--file", "-f",
        type=Path,
        default=DEFAULT_XLSX_PATH,
        help=f"Pfad zur Excel-Datei (Standard: {DEFAULT_XLSX_PATH})"
    )
    parser.add_argument(
        "--dry-run", "-n",
        action="store_true",
        help="Nur validieren, nichts in die DB schreiben"
    )
    parser.add_argument(
        "--update", "-u",
        action="store_true",
        help="Bestehende Ticker aktualisieren statt überspringen"
    )
    parser.add_argument(
        "--sheet",
        default="Sub-Industry ETFs",
        help="Name des Excel-Sheets (Standard: 'Sub-Industry ETFs')"
    )
    args = parser.parse_args()

    if args.dry_run:
        print("\n[DRY-RUN] Keine Änderungen an der Datenbank.\n")

    result = run_import(
        xlsx_path=args.file,
        dry_run=args.dry_run,
        update_existing=args.update,
        sheet_name=args.sheet,
    )

    print(result.summary())

    if result.messages:
        print("\nFehler & Warnungen:")
        for msg in result.messages:
            print(f"  {msg}")

    sys.exit(0 if result.errors == 0 else 1)


if __name__ == "__main__":
    main()
